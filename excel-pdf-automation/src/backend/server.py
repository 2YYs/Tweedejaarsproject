# -*- coding: utf-8 -*-
"""
Dit script fungeert als de backend-server voor de Excel-PDF-automaton.
Het bevat alle logica die oorspronkelijk in het zebi.ipynb notebook is ontwikkeld.
"""

# --- Imports ---
import os
import re
import json
import time
import uuid
import traceback
from datetime import datetime

import pandas as pd
import pdfplumber
import pytesseract
import requests
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pdf2image import convert_from_path
from openpyxl.styles import Font, Color
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
# PIL is een afhankelijkheid van pytesseract
from PIL import Image

# --- Configuratie ---
# De API-sleutel is hier hardcoded, zoals in het notebook.
# Voor productie is het beter om dit via een omgevingsvariabele te doen.
OPENROUTER_API_KEY = "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"
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL_NAME = "deepseek/deepseek-r1:free"
print("Configuratie geladen.")


# --- Functiedefinities ---

def extract_text_from_pdf_with_ocr(pdf_path, min_text_length=100):
    """
    Probeert eerst tekst digitaal te extraheren. Als dat te weinig tekst oplevert,
    schakelt het automatisch over naar OCR met Tesseract.
    """
    text = ""
    # Eerste poging: digitale extractie met pdfplumber
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\\n"
        print("Methode: Digitale extractie (pdfplumber) geslaagd.")
    except Exception as e:
        print(f"Fout tijdens digitale extractie: {e}. Probeert nu OCR.")
        text = ""  # Reset de tekst voor de OCR poging

    # Controleer of de digitale extractie genoeg tekst heeft opgeleverd
    if len(text.strip()) >= min_text_length:
        return text

    # Tweede poging: OCR met Tesseract voor ingescande PDF's
    print(f"Digitale extractie leverde < {min_text_length} tekens op. Overschakelen naar OCR...")
    try:
        images = convert_from_path(pdf_path)
        ocr_text = ""
        for i, image in enumerate(images):
            print(f"  Verwerken van pagina {i+1} met OCR...")
            ocr_text += pytesseract.image_to_string(image, lang='nld') + "\\n"
        print("Methode: OCR-extractie (Tesseract) geslaagd.")
        return ocr_text
    except Exception as e:
        print(f"Fout tijdens OCR-extractie: {e}")
        return None

def extract_percentage_sentences(text):
    """
    Zoekt naar zinnen in de tekst die percentages bevatten,
    met een focus op contexten die gerelateerd kunnen zijn aan loonstijgingen.
    """
    if not text:
        return []
    
    # Regex om zinnen te vinden die relevante sleutelwoorden en een percentage bevatten.
    percentage_pattern = re.compile(
        r'([^.?!]*?(?:loon|salaris|cao|verhoging|stijging|toeslag)\\s*[^.?!]*?\\d[\\d.,]*\\s?%\\s*[^.?!]*?[.?!])',
        re.IGNORECASE
    )
    sentences_with_percentage = percentage_pattern.findall(text)

    # Verwijder overtollige spaties en newlines.
    clean_sentences = [s.strip().replace('\\n', ' ') for s in sentences_with_percentage]
    return clean_sentences

def classify_with_deepseek(sentence, api_key, api_url, model_name, max_retries=3, delay=5):
    """
    Roept de API aan om een zin te analyseren en loonstijgingen te CATEGORISEREN.
    """
    if not api_key:
        print("Fout: OpenRouter API-sleutel is leeg of niet ingesteld.")
        return None

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://localhost/sakkal-cao-analyzer",
        "X-Title": "Team Sakkal CAO Analyzer"
    }
    
    system_prompt = """Je bent een expert in het analyseren van Nederlandse CAO-teksten. Je taak is om **alleen** concrete, definitieve loonstijgingen te vinden, te extraheren en te categoriseren.

**ZEER BELANGRIJKE REGELS OM TE VOLGEN:**
- **NEGEER VOLLEDIG** alle zinnen die onderdeel zijn van een voorbeeld, berekening of hypothese.
- Zoek naar sleutelwoorden zoals: **"Voorbeeld:", "Rekenvoorbeeld", "Stel dat", "Als ... dan", "Berekening"**. Als je zo'n sleutelwoord ziet, is de zin bijna altijd een voorbeeld en moet je een lege `verhogingen` lijst teruggeven.
- Negeer ook verhogingen die voorwaardelijk zijn ('tenzij', 'indien') of die alleen een vergelijking maken met het verleden.
- Focus **alleen** op definitieve, vastgelegde collectieve loonsverhogingen voor de toekomst.

**Jouw taak:**
Analyseer de zin en geef een JSON-object terug. Het object moet een lijst `verhogingen` bevatten. Voor **elke** gevonden loonstijging, maak een object met VIER sleutels:
1. `datum`: De ingangsdatum (formaat: "DD/MM/YYYY").
2. `percentage`: Het percentage (als een getal, bv. 3.5).
3. `categorie`: Classificeer het type verhoging. Kies uit:
    - "standaard": Een algemene, collectieve loonsverhoging.
    - "verlofdag_omzetting": Een verhoging die voortkomt uit het inruilen van verlofdagen.
    - "dienstjaren_toeslag": Een verhoging voor medewerkers met een bepaald aantal dienstjaren.
    - "WML_koppeling": Een verhoging die direct gekoppeld is aan het Wettelijk Minimumloon (WML).
    - "anders": Een andere, specifieke maar gegarandeerde verhoging.
4. `uitleg`: Een korte toelichting op je keuze.

**BELANGRIJK:** Je antwoord MOET **uitsluitend** een enkel, geldig JSON-object zijn met de sleutel 'verhogingen'. Geen extra tekst."""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": "De salarissen stijgen met 2% op 01-01-2025 en met nog eens 3% op 01-07-2025."},
        {"role": "assistant", "content": '''{"verhogingen": [{"datum": "01/01/2025", "percentage": 2.0, "categorie": "standaard", "uitleg": "Een standaard collectieve verhoging."}, {"datum": "01/07/2025", "percentage": 3.0, "categorie": "standaard", "uitleg": "Een tweede standaard collectieve verhoging."}]}'''},
        {"role": "user", "content": "Met ingang van 1 januari 2025 worden deze verlofdagen omgezet in een structurele salarisverhoging van 0,84%."},
        {"role": "assistant", "content": '''{"verhogingen": [{"datum": "01/01/2025", "percentage": 0.84, "categorie": "verlofdag_omzetting", "uitleg": "De verhoging is een directe omzetting van verlofdagen."}]}'''},
        {"role": "user", "content": "Het minimum bruto uursalaris van de Klantadviseur Level 1 per 1 januari 2026 wordt berekend door op het wettelijk minimumloon van 1 januari 2026 een percentage van 3,85% te tellen."},
        {"role": "assistant", "content": '''{"verhogingen": [{"datum": "01/01/2026", "percentage": 3.85, "categorie": "WML_koppeling", "uitleg": "De verhoging is direct gekoppeld aan het WML."}]}'''},
        {"role": "user", "content": "Voorbeeld: Als het bruto uursalaris volgens de cao op 1 januari 2026 met 3,88% stijgt en de Klantadviseur op 1 juli 2025 een beoordelingsverhoging van 0,80% heeft ontvangen, dan zal het uursalaris..."},
        {"role": "assistant", "content": '''{"verhogingen": [], "uitleg": "De zin begint met 'Voorbeeld:' en bevat 'Als...dan', wat duidt op een hypothetische berekening en geen definitieve loonafspraak."}'''},
        {"role": "user", "content": sentence}
    ]

    data = { "model": model_name, "messages": messages, "response_format": { "type": "json_object" }, "max_tokens": 500, "temperature": 0.0 }
    
    for attempt in range(max_retries):
        try:
            response = requests.post(api_url, json=data, headers=headers, timeout=45)
            response.raise_for_status()
            content = response.text
            start_index = content.find('{')
            end_index = content.rfind('}')
            if start_index != -1 and end_index != -1:
                json_str = content[start_index : end_index + 1]
                full_response_json = json.loads(json_str)
                message_content_str = full_response_json['choices'][0]['message']['content']
                return json.loads(message_content_str)
            return None
        except Exception as e:
            print(f"Fout tijdens poging {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(delay)
    return None

def analyze_pdfs(pdf_file_paths):
    """
    Orchestreert het volledige analyseproces voor een lijst van PDF-bestanden.
    """
    final_json_output = {}
    for pdf_path in pdf_file_paths:
        pdf_filename = os.path.basename(pdf_path)
        print(f'--- Start analyse voor: {pdf_filename} ---')
        extracted_text = extract_text_from_pdf_with_ocr(pdf_path)
        if not extracted_text:
            final_json_output[pdf_filename] = {"error": "Tekstextractie mislukt", "verhogingen": []}
            continue
        sentences = extract_percentage_sentences(extracted_text)
        if not sentences:
            final_json_output[pdf_filename] = {"error": "Geen relevante zinnen gevonden", "verhogingen": []}
            continue

        all_found_increases = []
        for sentence in sentences:
            result_json = classify_with_deepseek(sentence, OPENROUTER_API_KEY, OPENROUTER_API_URL, MODEL_NAME)
            if result_json and 'verhogingen' in result_json:
                all_found_increases.extend(result_json['verhogingen'])
            time.sleep(1)

        def sort_key(item):
            try:
                return datetime.strptime(item.get('datum', ''), '%d/%m/%Y')
            except (ValueError, TypeError):
                return datetime.max
        all_found_increases.sort(key=sort_key)
        
        final_json_output[pdf_filename] = {"verhogingen": all_found_increases}
        print(f'--- Analyse voor {pdf_filename} voltooid ---')
    return final_json_output

def create_excel_summary(analysis_results, output_filename):
    """
    Maakt een Excel-bestand van de analyseresultaten. Verhogingen op dezelfde
    datum worden in één cel gegroepeerd, met individueel gekleurde tekst per percentage.
    """
    # Map categorieën naar FONT-objecten voor Rich Text
    FONT_MAP = {
        "verlofdag_omzetting": Font(color=Color(rgb="FFC000")),  # Oranje/Goud
        "dienstjaren_toeslag": Font(color=Color(rgb="0070C0")), # Blauw
        "WML_koppeling": Font(color=Color(rgb="00B050")),        # Groen
        "anders": Font(color=Color(rgb="7030A0")),               # Paars
        "standaard": Font(color=Color(rgb="000000"))             # Zwart (standaard)
    }
    DEFAULT_FONT = FONT_MAP["standaard"]

    LEGEND_DESCRIPTIONS = {
        "standaard": "Standaard collectieve loonsverhoging.",
        "WML_koppeling": "Verhoging gekoppeld aan het Wettelijk Minimumloon.",
        "verlofdag_omzetting": "Verhoging door omzetting van verlofdagen.",
        "dienstjaren_toeslag": "Toeslag gebaseerd op aantal dienstjaren.",
        "anders": "Andere gespecificeerde verhoging."
    }

    # Stap 1: Groepeer verhogingen per datum voor elk bestand
    grouped_results = {}
    max_dates = 0
    for filename, data in analysis_results.items():
        increases_by_date = {}
        for verhoging in data.get('verhogingen', []):
            datum = verhoging.get('datum')
            if datum:
                # Robuuste datumconversie
                try:
                    parsed_date = datetime.strptime(datum, '%d/%m/%Y')
                    date_key = parsed_date.strftime('%d/%m/%Y')
                    if date_key not in increases_by_date:
                        increases_by_date[date_key] = []
                    increases_by_date[date_key].append(verhoging)
                except ValueError:
                    print(f"Ongeldig datumformaat overgeslagen: {datum} in {filename}")
        
        # Sorteer datums correct voor de output
        sorted_dates = sorted(increases_by_date.keys(), key=lambda d: datetime.strptime(d, '%d/%m/%Y'))
        grouped_results[filename] = [(d, increases_by_date[d]) for d in sorted_dates]
        if len(sorted_dates) > max_dates:
            max_dates = len(sorted_dates)

    # Stap 2: Bereid data voor in een platte structuur voor de DataFrame
    processed_data = []
    for filename, date_groups in grouped_results.items():
        row_data = {'Bestandsnaam': filename}
        for i, (date, increases) in enumerate(date_groups):
            num = i + 1
            row_data[f'{num}e datum'] = date
            # Plain text placeholder; Rich Text wordt later toegepast
            row_data[f'{num}e percentages'] = " / ".join(
                [f"{inc.get('percentage', 0):.2f}%".replace('.', ',') for inc in increases if isinstance(inc.get('percentage'), (int, float))]
            )
        processed_data.append(row_data)

    # Stap 3: Creëer de DataFrame
    df = pd.DataFrame()
    if processed_data:
        column_headers = ['Bestandsnaam']
        for i in range(1, max_dates + 1):
            column_headers.append(f'{i}e datum')
            column_headers.append(f'{i}e percentages')
        df = pd.DataFrame(processed_data, columns=column_headers).fillna('')

    # Stap 4: Schrijf naar Excel en pas Rich Text opmaak toe
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Samenvatting')
        worksheet = writer.sheets['Samenvatting']
        header_map = {cell.value: i + 1 for i, cell in enumerate(worksheet[1])}

        # Itereer opnieuw om de cellen te vullen met Rich Text
        for row_idx, filename in enumerate(df['Bestandsnaam'], start=2):
            if filename not in grouped_results:
                continue
            date_groups = grouped_results.get(filename, [])
            for i, (date, increases) in enumerate(date_groups):
                num = i + 1
                percentages_col_name = f'{num}e percentages'
                col_idx = header_map.get(percentages_col_name)
                if not col_idx: continue

                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                rich_text_payload = []
                for j, increase in enumerate(increases):
                    # Alleen verwerken als het een geldig percentage is
                    if not isinstance(increase.get('percentage'), (int, float)):
                        continue
                        
                    if j > 0 and rich_text_payload: # Zorg dat we niet starten met een separator
                        rich_text_payload.append(TextBlock(DEFAULT_FONT, " / "))
                    
                    percentage_str = f"{increase.get('percentage', 0):.2f}%".replace('.', ',')
                    font = FONT_MAP.get(increase.get('categorie', 'standaard'), DEFAULT_FONT)
                    rich_text_payload.append(TextBlock(font, percentage_str))
                
                if rich_text_payload:
                    cell.value = CellRichText(rich_text_payload)
                else:
                    cell.value = "" # Leegmaken als er geen geldige percentages waren

        # Stap 5: Pas kolombreedtes aan en voeg de legenda toe
        def get_cell_text_length(cell):
            if isinstance(cell.value, CellRichText):
                return len("".join(block.text for block in cell.value))
            return len(str(cell.value))

        for i, column_cells in enumerate(worksheet.columns):
            if i < len(df.columns):
                length = max(get_cell_text_length(cell) for cell in column_cells)
                column_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[column_letter].width = length + 4

        if not df.empty:
            legend_start_col = len(df.columns) + 2
            
            header_cell = worksheet.cell(row=1, column=legend_start_col, value="LEGENDA")
            header_cell.font = Font(bold=True)
            
            current_row = 2
            for category, description in LEGEND_DESCRIPTIONS.items():
                cell = worksheet.cell(row=current_row, column=legend_start_col, value=description)
                cell.font = FONT_MAP.get(category, DEFAULT_FONT)
                current_row += 1
            
            legend_col_letter = get_column_letter(legend_start_col)
            max_len = max(len(d) for d in LEGEND_DESCRIPTIONS.values()) if LEGEND_DESCRIPTIONS else 10
            worksheet.column_dimensions[legend_col_letter].width = max(max_len, len("LEGENDA")) + 4

    print(f'Excel-bestand succesvol aangemaakt op: {output_filename}')


# --- Flask Webserver ---
app = Flask(__name__)
CORS(app)
UPLOAD_FOLDER = 'temp_uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/api/process', methods=['POST'])
def process_uploaded_pdfs():
    print("\\n[SERVER LOG] Received new request for /api/process")
    if 'files' not in request.files:
        return jsonify({"error": "Geen bestanden meegegeven"}), 400
    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({"error": "Geen bestanden geselecteerd"}), 400
    
    request_id = str(uuid.uuid4())
    request_upload_dir = os.path.join(UPLOAD_FOLDER, request_id)
    os.makedirs(request_upload_dir, exist_ok=True)
    
    saved_file_paths = []
    for file in files:
        filename = secure_filename(file.filename)
        filepath = os.path.join(request_upload_dir, filename)
        file.save(filepath)
        saved_file_paths.append(filepath)
        
    try:
        analysis_results = analyze_pdfs(saved_file_paths)
        if not analysis_results or all(not v.get('verhogingen') for v in analysis_results.values()):
            return jsonify({"error": "Analyse heeft geen bruikbare loonsverhogingen opgeleverd."}), 500
            
        output_excel_filename = f"cao_samenvatting_{request_id}.xlsx"
        output_excel_path = os.path.join(request_upload_dir, output_excel_filename)
        create_excel_summary(analysis_results, output_excel_path)
        
        return send_file(output_excel_path, as_attachment=True, download_name='cao_samenvatting.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Onverwachte serverfout: {str(e)}"}), 500

# --- Server Start ---
if __name__ == '__main__':
    print('--- Flask server starten vanaf server.py ---')
    print('Server is live op http://127.0.0.1:5001')
    app.run(host='127.0.0.1', port=5001, debug=True) 